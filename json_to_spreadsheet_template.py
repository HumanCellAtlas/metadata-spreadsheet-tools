import logging
from optparse import OptionParser

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

# hard coded tab ordering
tab_ordering = ["project", "project.publications", "contact", "donor_organism", "familial_relationship",
                "specimen_from_organism", "cell_suspension",
                "cell_line", "cell_line.publications", "organoid", "collection_process", "dissociation_process",
                "enrichment_process", "library_preparation_process",
                "sequencing_process", "purchased_reagents", "protocol", "sequence_file"]


class SpreadsheetCreator:

    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def generate_spreadsheet(self, schema_base_uri, schema_types, schema_modules, output):
        values = {}
        try:
            print("creating spreadsheet from:\t" + schema_base_uri)

            for index, schema_module in enumerate(schema_modules):
                schema_modules[index] = schema_base_uri + schema_module

            # for each schema, gather the values for the relevant tab(s)
            for schema_type in schema_types:
                v = self._gather_values(schema_base_uri + schema_type, schema_modules)
                values.update(v)
            # Build the spreadsheet from the retrieved values
            self._build_spreadsheet(values, output)
        except ValueError as e:
            self.logger.error("Error:" + str(e))
            raise e

    def _gather_values(self, schema_uri, schema_modules):
        # get the schema of HTTP
        req = requests.get(schema_uri)
        print("- collecting fields from:\t" + schema_uri)
        # if the schema is successfully retrieved, process it, else return an error message
        if req.status_code == requests.codes.ok:
            json_raw = req.json()
            entities = {}
            entity_title = json_raw["title"]
            properties = json_raw["properties"]
            values = []

            for prop in properties:
                # if a property has an array of references (potential 1-to-many relationship)
                if "items" in properties[prop] and "$ref" in properties[prop]["items"]:
                    self._add_fields_from_array_of_schemas(entities, entity_title, prop, properties, schema_modules,
                                                           values)
                # if a property does not include a user_friendly tag but includes a reference
                elif "$ref" in properties[prop]:
                    self._add_fields_from_referenced_schema(prop, properties, schema_modules, values)
                # if a property has a user_friendly tag, include it as a direct field.
                # This includes ontology module references as these should not be exposed to users
                elif "user_friendly" in properties[prop]:
                    self._add_field_directly(prop, properties, values)

            self._add_relationship_fields(schema_uri, values)

            entities[entity_title] = values
            return entities

        else:
            self.logger.error(schema_uri + " does not exist")

    @staticmethod
    def _add_field_directly(prop, properties, values):
        print("\t\tadding " + properties[prop]["user_friendly"])
        description = None
        example = None
        if "description" in properties[prop]:
            description = properties[prop]["description"]
        if "example" in properties[prop]:
            example = "e.g. " + str(properties[prop]["example"])
        values.append({"header": properties[prop]["user_friendly"], "description": description,
                       "example": example})

    # gather the properties for the references and format them to become
    # their own spreadsheet tab
    def _add_fields_from_array_of_schemas(self, entities, entity_title, prop, properties, schema_modules, values):
        module = properties[prop]["items"]["$ref"]
        if "ontology" not in module and module in schema_modules:
            module_values = self._gather_values(module, None)
            # add primary entity ID to cross reference with main entity
            for primary in values:
                if "ID" in primary["header"]:
                    for key in module_values.keys():
                        t = primary["header"]
                        t = t.replace(" ID", "").lower()
                        d = "ID for " + t + " this " + key + " relates to"
                        module_values[key].append({"header": t,
                                                   "description": d,
                                                   "example": None})
                    break

            # special name cases for publication tabs
            if entity_title == "project" and "publication" in module_values.keys():
                module_values["project.publications"] = module_values.pop("publication")
            if entity_title == "cell_line" and "publication" in module_values.keys():
                module_values["cell_line.publications"] = module_values.pop("publication")
            entities.update(module_values)

    # fetch the contents of that reference and add them directly to the properties for this sheet
    def _add_fields_from_referenced_schema(self, prop, properties, schema_modules, values):
        module = properties[prop]["$ref"]
        if "ontology" not in module and ("_core" in module or module in schema_modules):
            module_values = self._gather_values(module, schema_modules)
            print("\t- adding fields from $ref:\t" + module)
            for key in module_values.keys():
                # special case for naming UMI barcodes
                if prop == "umi_barcode":
                    for entry in module_values[key]:
                        entry["header"] = "UMI " + entry["header"]
                # special case for naming cell barcodes
                if prop == "cell_barcode":
                    for entry in module_values[key]:
                        entry["header"] = "Cell " + entry["header"]

                values.extend(module_values[key])

    @staticmethod
    def _add_relationship_fields(schema_uri, values):
        if "type/biomaterial" in schema_uri:
            values.append(
                {"header": "Process IDs", "description": "IDs of processes for which this biomaterial is an input",
                 "example": None})
        if "type/process" in schema_uri:
            values.append(
                {"header": "Protocol IDs", "description": "IDs of protocols which this process implements",
                 "example": None})
        if "type/file" in schema_uri:
            values.append(
                {"header": "Biomaterial ID", "description": "ID of the biomaterial to which this file relates",
                 "example": None})
            values.append(
                {"header": "Sequencing process ID",
                 "description": "ID of the sequencing process to which this file relates",
                 "example": None})

    def _build_spreadsheet(self, values, output_location):
        wb = Workbook()

        # for each tab entry in the values dictionary, create a new worksheet
        # for tab_name in values.keys():
        for tab_name in tab_ordering:
            if tab_name in values.keys():
                headers = values[tab_name]

                ws = wb.create_sheet(title=tab_name)
                col = 1

                # Optional set of descriptors what each of the 3 top rows contains
                # put each description in row 1, example in row 2 and header in row 3, then increment the column index
                for header in headers:
                    ws.cell(column=col, row=1, value=header["header"]).font = Font(bold=True)
                    ws.cell(column=col, row=1, value=header["header"]).fill = PatternFill("solid",  fgColor="D9D9D9")
                    ws.cell(column=col, row=2, value=header["description"]).font = Font(italic=True, color="595959")
                    ws.cell(column=col, row=3, value=header["example"]).font = Font(color="595959")
                    col += 1

        # remove the blank worksheet that is automatically created with the spreadsheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        print("\ncreated spreadsheet with the following tabs:")
        for sheetname in wb.sheetnames:
            print('- ' + sheetname)
            self._autosize_columns(wb[sheetname])
        wb.save(filename=output_location)

    @staticmethod
    def _autosize_columns(worksheet):
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width


if __name__ == '__main__':
    parser = OptionParser()
    parser.add_option("-s", "--schema", dest="schema_uri",
                      help="Base schema URI for the metadata")
    parser.add_option("-o", "--output", dest="output",
                      help="Output directory and file where to save the template spreadsheet", metavar="FILE")
    parser.add_option("-t", "--types", dest="schema_types",
                      help="Schema types to include in the spreadsheet")
    parser.add_option("-i", "--include", dest="include",
                      help="Schema modules to include in the spreadsheet")

    (options, args) = parser.parse_args()

    # if not options.schema_uri:
    #     print ("You must supply a base schema URI for the metadata")
    #     exit(2)

    provided_schema_types = options.schema_types.split(",")
    dependencies = options.include.split(",")

    # for index, dependency in enumerate(dependencies):
    #    dependencies[index] = options.schema_uri + dependency

    generator = SpreadsheetCreator()
    generator.generate_spreadsheet(options.schema_uri, provided_schema_types, dependencies, options.output)
